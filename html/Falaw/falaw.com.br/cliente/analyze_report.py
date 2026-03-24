"""
analyze_report.py — Falaw Advogados
====================================
Lê um relatório Excel (.xlsx) e extrai os dados de processos trabalhistas
para exibição automática no dashboard do portal do cliente.

USO:
    python analyze_report.py caminho/do/relatorio.xlsx
    python analyze_report.py caminho/do/relatorio.xlsx --client-id abc123 --report-id xyz789

SAÍDA:
    Gera/atualiza o arquivo `data/stats.json` com os dados extraídos.
    O dashboard do cliente carrega esse arquivo automaticamente.

DEPENDÊNCIAS:
    pip install openpyxl pandas
"""

import sys
import json
import os
import argparse
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
except ImportError:
    print("ERRO: Instale as dependências:\n  pip install openpyxl pandas")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# CONFIGURAÇÃO DE COLUNAS
# Ajuste os nomes abaixo para bater com as colunas do seu Excel.
# Pode ser o cabeçalho exato da coluna OU uma lista de alternativas.
# ─────────────────────────────────────────────────────────────

COLUMN_MAP = {
    # Total de processos ativos no período
    "processos": [
        "total de processos", "processos ativos", "total processos",
        "quantidade de processos", "qtd processos", "processos"
    ],

    # Decisões favoráveis
    "favoraveis": [
        "decisões favoráveis", "decisoes favoraveis", "favoráveis",
        "favoraveis", "ganhos", "favorável", "procedente",
        "procedentes", "favorável ao cliente", "vitórias"
    ],

    # Decisões desfavoráveis
    "desfavoraveis": [
        "decisões desfavoráveis", "decisoes desfavoraveis",
        "desfavoráveis", "desfavoraveis", "perdas", "improcedente",
        "improcedentes", "desfavorável ao cliente"
    ],

    # Novas ações distribuídas no período
    "novas": [
        "novas ações", "novas acoes", "novos processos",
        "distribuídos", "distribuidos", "novas", "ingressos"
    ],

    # Processos encerrados no período
    "encerrados": [
        "encerrados", "arquivados", "finalizados", "extintos",
        "processos encerrados", "baixados"
    ],

    # Processos em andamento
    "andamento": [
        "em andamento", "andamento", "ativos", "em curso",
        "pendentes", "aguardando", "em tramitação"
    ],
}

# Se o Excel tiver uma linha de resumo/totais, defina o rótulo dela aqui
SUMMARY_ROW_LABELS = [
    "total", "totais", "resumo", "subtotal", "geral", "consolidado"
]


# ─────────────────────────────────────────────────────────────
# FUNÇÕES AUXILIARES
# ─────────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    """Normaliza string para comparação (minúsculas, sem acento)."""
    if not isinstance(text, str):
        return ""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", text.lower().strip())
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def find_column(df_columns, aliases: list[str]) -> str | None:
    """Encontra a coluna do DataFrame que bate com algum dos aliases."""
    norm_cols = {normalize(c): c for c in df_columns}
    for alias in aliases:
        key = normalize(alias)
        if key in norm_cols:
            return norm_cols[key]
    return None


def extract_number(value) -> int | None:
    """Converte um valor de célula para inteiro."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def read_stats_from_sheet(df: pd.DataFrame) -> dict:
    """
    Tenta extrair os stats de um DataFrame.
    Estratégias:
      1. Procura colunas com os nomes mapeados → pega a soma ou última linha de dados
      2. Procura uma linha de "TOTAL" e lê os valores dessa linha
    """
    stats = {}
    df.columns = [str(c).strip() for c in df.columns]

    # Estratégia 1: colunas nomeadas
    for stat_key, aliases in COLUMN_MAP.items():
        col_name = find_column(df.columns, aliases)
        if col_name:
            series = df[col_name].dropna()
            if len(series) > 0:
                # Se tiver uma linha de total/resumo, usa ela
                for idx, val in df.iterrows():
                    row_label = normalize(str(val.iloc[0]))
                    if any(lbl in row_label for lbl in SUMMARY_ROW_LABELS):
                        n = extract_number(val[col_name])
                        if n is not None:
                            stats[stat_key] = n
                            break
                # Caso contrário usa a soma
                if stat_key not in stats:
                    total = series.apply(extract_number).dropna().sum()
                    if total > 0:
                        stats[stat_key] = int(total)

    return stats


def extract_from_excel(filepath: str) -> dict:
    """
    Lê o arquivo Excel e extrai os dados de todas as abas.
    Retorna um dict com os stats consolidados.
    """
    print(f"\n📂 Abrindo arquivo: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"   Abas encontradas: {wb.sheetnames}")

    all_stats = {}

    for sheet_name in wb.sheetnames:
        print(f"\n   📋 Analisando aba: '{sheet_name}'")
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
            print(f"      Colunas: {list(df.columns)}")
            sheet_stats = read_stats_from_sheet(df)
            if sheet_stats:
                print(f"      ✅ Dados extraídos: {sheet_stats}")
                # Mescla — soma valores de múltiplas abas se necessário
                for k, v in sheet_stats.items():
                    if k not in all_stats:
                        all_stats[k] = v
            else:
                print("      ⚠️  Nenhum dado reconhecido nesta aba.")
        except Exception as e:
            print(f"      ❌ Erro ao ler aba: {e}")

    return all_stats


# ─────────────────────────────────────────────────────────────
# SAÍDA JSON
# ─────────────────────────────────────────────────────────────

def save_stats(stats: dict, filepath_xlsx: str, client_id: str = None,
               report_id: str = None, period: str = None):
    """
    Salva os stats extraídos em data/stats.json
    no formato esperado pelo dashboard.
    """
    output_dir = Path(__file__).parent / "data"
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "stats.json"

    # Lê JSON existente para não sobrescrever outros clientes
    existing = {}
    if output_file.exists():
        try:
            with open(output_file, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass

    entry = {
        "generated": datetime.now().isoformat(),
        "sourceFile": Path(filepath_xlsx).name,
        "period": period or "",
        "clientId": client_id or "",
        "reportId": report_id or "",
        "stats": {
            "processos":     stats.get("processos"),
            "favoraveis":    stats.get("favoraveis"),
            "desfavoraveis": stats.get("desfavoraveis"),
            "novas":         stats.get("novas"),
            "encerrados":    stats.get("encerrados"),
            "andamento":     stats.get("andamento"),
        }
    }

    # Chave única: usa reportId, ou período+arquivo, ou só o nome do arquivo
    if client_id and report_id:
        key = f"{client_id}_{report_id}"
    elif period:
        key = f"{period}_{Path(filepath_xlsx).stem}"
    else:
        key = Path(filepath_xlsx).stem

    existing[key] = entry
    # "latest" sempre aponta para o relatório mais recente processado
    # O dashboard usa isso para atualizar KPIs e gráficos automaticamente
    existing["latest"] = entry

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Stats salvos em: {output_file}")
    print(json.dumps(entry["stats"], ensure_ascii=False, indent=2))
    return output_file


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extrai dados de processos de um relatório Excel para o dashboard Falaw."
    )
    parser.add_argument("arquivo", help="Caminho do arquivo .xlsx")
    parser.add_argument("--client-id",  default="", help="ID do cliente no portal")
    parser.add_argument("--report-id",  default="", help="ID do relatório no portal")
    parser.add_argument("--period",     default="", help="Período do relatório (ex: Março/2026)")
    args = parser.parse_args()

    if not os.path.exists(args.arquivo):
        print(f"ERRO: Arquivo não encontrado: {args.arquivo}")
        sys.exit(1)

    stats = extract_from_excel(args.arquivo)

    if not stats:
        print("\n⚠️  ATENÇÃO: Não foi possível extrair dados automaticamente.")
        print("   Verifique se os nomes das colunas no Excel batem com os mapeados em COLUMN_MAP.")
        print("   Colunas esperadas (exemplos):")
        for k, aliases in COLUMN_MAP.items():
            print(f"     {k}: {aliases[:3]}")
        sys.exit(1)

    save_stats(stats, args.arquivo, args.client_id, args.report_id, args.period)
    print("\n🚀 Pronto! Recarregue o dashboard para ver os dados atualizados.")


if __name__ == "__main__":
    main()
