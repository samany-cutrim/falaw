#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
processar_cliente.py — Debug e extração de dados do Excel de cliente (Falaw Advogados)

Uso:
    python processar_cliente.py "caminho/para/RELATORIO_CLIENTE.xlsx"

O script imprime:
  1. Abas encontradas
  2. Dados da aba DASHBOARD (KPIs: ativos, encerrados, exposição, taxa de êxito)
  3. Nome da aba de processos ativos encontrada e total de linhas
  4. Mapeamento de colunas detectado (cabeçalho + índices)
  5. Amostra de 5 registros parseados
  6. Alertas se colunas importantes não forem encontradas
"""

import sys
import os
import re
import unicodedata
from collections import Counter

try:
    import openpyxl
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False

try:
    import pandas as pd
    USE_PANDAS = True
except ImportError:
    USE_PANDAS = False

if not USE_PANDAS and not USE_OPENPYXL:
    print("ERRO: instale pandas+openpyxl:  pip install pandas openpyxl")
    sys.exit(1)


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def norm(s):
    """Normaliza string: maiúsculo, sem acentos, sem espaço extra."""
    if s is None:
        return ''
    s = str(s).strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

def parse_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[R$\s]', '', str(v)).replace('.', '').replace(',', '.').replace('%', '')
    try:
        return float(s)
    except ValueError:
        return None

def find_sheet(sheet_names, *patterns):
    """Retorna o primeiro nome de aba cujo nome normalizado contém qualquer padrão."""
    for name in sheet_names:
        nn = norm(name)
        for p in patterns:
            if p in nn:
                return name
    return None


# ─── LEITURA DO WORKBOOK ───────────────────────────────────────────────────────

def load_workbook_rows(path, sheet_name):
    """Lê uma aba e retorna lista de listas (linhas × colunas)."""
    if USE_PANDAS:
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
        return df.fillna('').values.tolist()
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet_name]
        return [[str(cell.value) if cell.value is not None else '' for cell in row] for row in ws.iter_rows()]

def get_sheet_names(path):
    if USE_PANDAS:
        xl = pd.ExcelFile(path)
        return xl.sheet_names
    else:
        wb = openpyxl.load_workbook(path, read_only=True)
        return wb.sheetnames


# ─── PARSE DA ABA DASHBOARD ───────────────────────────────────────────────────

def parse_dashboard_tab(path, sheet_name):
    rows = load_workbook_rows(path, sheet_name)
    kpis = {'ativos': None, 'encerrados': None, 'exposicao': None, 'taxaExito': None, 'periodo': ''}

    for i, row in enumerate(rows[:15]):
        for v in row:
            if 'RELAT' in norm(v) and 'MENSAL' in norm(v):
                kpis['periodo'] = str(v).strip()
                break

    for i, row in enumerate(rows[:20]):
        labels = [norm(v) for v in row]
        if any('PROCESSOS ATIVOS' in l for l in labels):
            val_row = rows[i + 1] if i + 1 < len(rows) else []
            for ci, label in enumerate(labels):
                v = val_row[ci] if ci < len(val_row) else None
                if 'PROCESSOS ATIVOS' in label and 'ENCERRADO' not in label:
                    kpis['ativos'] = parse_num(v)
                elif 'EXPOSICAO' in label or 'EXPOSIÇÃO' in label:
                    kpis['exposicao'] = parse_num(v)
                elif 'ENCERRADO' in label and 'PROCESSOS' in label:
                    kpis['encerrados'] = parse_num(v)
                elif 'TAXA' in label and ('EXITO' in label or 'EXITO' in label):
                    kpis['taxaExito'] = parse_num(v)
            break

    return kpis


# ─── DETECÇÃO DE COLUNA ───────────────────────────────────────────────────────

COLUMN_PATTERNS = {
    'num_processo':  ['PROCESSO', 'NUMERO', 'NR', 'NPROC', 'NO PROC'],
    'trt':           ['TRT', 'ORGAO', 'ORGÃO'],
    'data_distrib':  ['DATA DA DISTRIBUICAO', 'DATA DE DISTRIBUICAO', 'DATA DE INSTAURACAO',
                      'DATA DO AJUIZAMENTO', 'DATA DE RECEBIMENTO', 'DATA DISTRIBUICAO'],
    'reclamante':    ['RECLAMANTE', 'POLO ATIVO', 'PARTE CONTRARIA', 'PARTES CONTRARIAS'],
    'reclamada':     ['RECLAMADA', 'PARTE PRINCIPAL', 'POLO PASSIVO', 'AGENTE'],
    'risco':         ['RISCO DE PERDA', 'RISCO'],
    'status':        ['STATUS ATUAL', 'STATUS MENSAL', 'STATUS', 'PRINCIPAIS ANDAMENTOS'],
    'fase':          ['FASE', 'INSTANCIA', 'INSTÂNCIA'],
    'uf':            ['UF', 'CIDADE/UF', 'ESTADO', 'COMARCA'],
    'vara':          ['VARA', 'JUIZO', 'JUÍZO'],
    'juiz':          ['JUIZ', 'JUIZA', 'MAGISTRADO', 'MAGISTRADA', 'RELATOR', 'RELATORA',
                      'DESEMBARGADOR', 'MINISTRO'],
    'adv':           ['ADV. CONTRARIO', 'ADV CONTRARIO', 'ESCRITORIO DA PARTE CONTRARIA',
                      'ESCRITORIO', 'ADVOGADO CONTRARIO'],
    'responsabilidade': ['RESPONSABILIDADE', 'PAPEL DO AGENTE', 'PAPEL', 'TIPO DE RESPONSABILIDADE'],
    'sentenca':      ['SENTENCA', 'RESULTADO', 'DESFECHO', 'LIMINAR'],
}

def find_header_row(rows, max_scan=20):
    """Encontra a linha de cabeçalho verificando quantas colunas-chave aparecem."""
    best = (-1, -1)
    for i, row in enumerate(rows[:max_scan]):
        normed = [norm(v) for v in row]
        score = sum(
            1 for pats in COLUMN_PATTERNS.values()
            if any(p in cell for p in pats for cell in normed)
        )
        if score > best[1]:
            best = (i, score)
    return best[0]

def map_columns(header_row):
    normed = [norm(v) for v in header_row]
    mapping = {}
    for field, patterns in COLUMN_PATTERNS.items():
        idx = -1
        for ci, cell in enumerate(normed):
            if any(p in cell for p in patterns):
                idx = ci
                break
        mapping[field] = idx
    return mapping


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python processar_cliente.py \"caminho/para/RELATORIO.xlsx\"")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"ERRO: arquivo não encontrado: {path}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  ANÁLISE: {os.path.basename(path)}")
    print(f"{'='*60}\n")

    sheet_names = get_sheet_names(path)
    print(f"[1] Abas encontradas ({len(sheet_names)}):")
    for i, n in enumerate(sheet_names):
        print(f"     {i+1}. {n}")
    print()

    # ── Dashboard tab ──────────────────────────────────────────────────────
    dash_sheet = find_sheet(sheet_names, 'DASHBOARD')
    if dash_sheet:
        print(f"[2] Aba DASHBOARD encontrada: '{dash_sheet}'")
        kpis = parse_dashboard_tab(path, dash_sheet)
        print(f"     Período    : {kpis['periodo'] or '(não encontrado)'}")
        print(f"     Ativos     : {kpis['ativos']}")
        print(f"     Encerrados : {kpis['encerrados']}")
        print(f"     Exposição  : {kpis['exposicao']}")
        print(f"     Taxa Êxito : {kpis['taxaExito']}")
        print()
        # Also dump first 10 rows of dashboard for manual inspection
        rows = load_workbook_rows(path, dash_sheet)
        print(f"     → Primeiras 10 linhas da aba '{dash_sheet}' (para inspeção manual):")
        for i, row in enumerate(rows[:10]):
            non_empty = [str(v) for v in row if str(v).strip()]
            if non_empty:
                print(f"       Linha {i+1}: {' | '.join(non_empty[:8])}")
        print()
    else:
        print("[2] Aba DASHBOARD: NÃO ENCONTRADA\n")

    # ── Active processes tab ───────────────────────────────────────────────
    ACTIVE_PATTERNS = ['ATIVO', 'JUDICIAL', 'TRABALHISTA', 'ANDAMENTO', 'PROCESSOS',
                       'ACOMPANHAMENTO', 'CARTEIRA']
    active_sheet = find_sheet(sheet_names, *ACTIVE_PATTERNS)
    if not active_sheet:
        # Try to find by process number pattern in first 20 rows
        for n in sheet_names:
            rows = load_workbook_rows(path, n)
            if any(re.search(r'\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}', str(v))
                   for row in rows[:20] for v in row):
                active_sheet = n
                break

    if active_sheet:
        print(f"[3] Aba de processos ativos: '{active_sheet}'")
        rows = load_workbook_rows(path, active_sheet)
        hdr_idx = find_header_row(rows)
        if hdr_idx < 0:
            print("     AVISO: nenhuma linha de cabeçalho reconhecível encontrada.")
        else:
            header = rows[hdr_idx]
            mapping = map_columns(header)
            data_rows = [r for r in rows[hdr_idx+1:] if any(str(v).strip() for v in r)]
            print(f"     Linha de cabeçalho: {hdr_idx + 1}")
            print(f"     Total de linhas de dados: {len(data_rows)}")
            print()
            print("[4] Mapeamento de colunas detectado:")
            missing = []
            for field, idx in mapping.items():
                if idx >= 0:
                    col_name = str(header[idx]).strip() if idx < len(header) else '?'
                    print(f"     {field:<20} → col {idx+1:>3}  ({col_name})")
                else:
                    missing.append(field)
            if missing:
                print()
                print(f"     ⚠ COLUNAS NÃO ENCONTRADAS: {', '.join(missing)}")
            print()
            print("[5] Amostra de registros (primeiros 5 com número de processo):")
            count = 0
            for row in data_rows:
                if count >= 5:
                    break
                proc_idx = mapping.get('num_processo', -1)
                proc = str(row[proc_idx]).strip() if proc_idx >= 0 and proc_idx < len(row) else ''
                if not re.search(r'\d{7,}', proc):
                    # Fallback: find in row
                    proc = next((str(v) for v in row if re.search(r'\d{7}-\d{2}\.\d{4}', str(v or ''))), '')
                if not proc:
                    continue
                sample = {f: (str(row[i]).strip() if i >= 0 and i < len(row) else '') for f, i in mapping.items() if i >= 0}
                print(f"     {count+1}. Processo: {proc}")
                for k, v in sample.items():
                    if v and k != 'num_processo':
                        print(f"        {k:<20}: {v[:60]}")
                print()
                count += 1

            # Responsabilidade distribution
            resp_idx = mapping.get('responsabilidade', -1)
            if resp_idx >= 0:
                resp_vals = [str(row[resp_idx]).strip() for row in data_rows if resp_idx < len(row) and str(row[resp_idx]).strip()]
                if resp_vals:
                    print("[6] Distribuição de Responsabilidade:")
                    for val, cnt in Counter(resp_vals).most_common(10):
                        print(f"     {cnt:>4}×  {val}")
                    print()

            # Juiz/relator sample
            juiz_idx = mapping.get('juiz', -1)
            if juiz_idx >= 0:
                juiz_vals = [str(row[juiz_idx]).strip() for row in data_rows
                             if juiz_idx < len(row) and str(row[juiz_idx]).strip()
                             and not re.match(r'^\d[\d.,/ -]*$', str(row[juiz_idx]).strip())]
                print(f"[7] Top 10 julgadores ({len(juiz_vals)} registros com nome):")
                for val, cnt in Counter(juiz_vals).most_common(10):
                    print(f"     {cnt:>4}×  {val}")
                print()
            else:
                print("[7] Coluna de julgador não encontrada — verifique o cabeçalho.\n")

    else:
        print("[3] Aba de processos ativos: NÃO ENCONTRADA")
        print("     Nomes disponíveis:")
        for n in sheet_names:
            print(f"       - {n}")
        print()

    print("Análise concluída.\n")


if __name__ == '__main__':
    main()
